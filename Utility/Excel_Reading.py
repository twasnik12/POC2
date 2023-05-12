import openpyxl

import Ellicium.DBConfigurations.Config_file

xl = Ellicium.DBConfigurations.Config_file.Excel_Report_File


def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def read_data(file, sheetname, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=columnno).value


def write_data(file, sheetname, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)




#write_data(xl, 'Sheet1', 2, 5, "TestCase_Failed")

# Update in ConfigFile
"""Excel_Report_File = r'C:\\Users\\Prith\\Desktop\\test_results.xlsx'
Sheet_Name = 'Sheet1'
"""
