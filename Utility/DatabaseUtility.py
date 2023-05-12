import openpyxl
import pyodbc

import Ellicium.DBConfigurations.Config_file


class DatabaseConnection:
    def Qurry1(self):

        cnxn_str = ("Driver={SQL Server};""Server=RLPT-0015\SQLEXPRESS;""Database=SuperStore;""Trusted_Connection=yes;")

        my_cnxn = pyodbc.connect(cnxn_str)

        cursor = my_cnxn.cursor()
        book = openpyxl.load_workbook(Ellicium.DBConfigurations.Config_file.Excel_Report_File)
        sheet = book.active
        cell = sheet.cell(row=2, column=3)

        cursor.execute(cell.value)
        db_list1 = []

        for i in cursor:
            db_list1.append(i)


        #print("list1",db_list1)

        return db_list1

    def Qurry2(self):
        cnxn_str = ("Driver={SQL Server};""Server=RLPT-0015\SQLEXPRESS;""Database=SuperStore;""Trusted_Connection=yes;")

        my_cnxn = pyodbc.connect(cnxn_str)

        cursor = my_cnxn.cursor()

        # cursor.execute('select Customer_Segment as'Customer Segment',Product_Category as 'Product Category',ROUND(sum(unit_Price),2) "Sum of Unit Price" from Superstoretestinggroup by Customer_Segment, Product_Categoryorder by Customer_Segment, Product_Category asc')

        book = openpyxl.load_workbook(Ellicium.DBConfigurations.Config_file.Excel_Report_File)
        sheet = book.active
        cell = sheet.cell(row=3, column=3)
        # print(cell.value)  # qurry

        cursor.execute(cell.value)
        db_list2 = []

        for i in cursor:
            db_list2.append(i)

        return db_list2

    def Qurry3(self):
        cnxn_str = ("Driver={SQL Server};""Server=RLPT-0015\SQLEXPRESS;""Database=SuperStore;""Trusted_Connection=yes;")

        my_cnxn = pyodbc.connect(cnxn_str)

        cursor = my_cnxn.cursor()

        # cursor.execute('select Customer_Segment as'Customer Segment',Product_Category as 'Product Category',ROUND(sum(unit_Price),2) "Sum of Unit Price" from Superstoretestinggroup by Customer_Segment, Product_Categoryorder by Customer_Segment, Product_Category asc')

        book = openpyxl.load_workbook(Ellicium.DBConfigurations.Config_file.Excel_Report_File)
        sheet = book.active
        cell = sheet.cell(row=4, column=3)
        # print(cell.value)  # qurry

        cursor.execute(cell.value)
        db_list3 = []

        for i in cursor:
            db_list3.append(i)

        return db_list3

    def Qurry4(self):

        cnxn_str = ("Driver={SQL Server};""Server=RLPT-0015\SQLEXPRESS;""Database=SuperStore;""Trusted_Connection=yes;")

        my_cnxn = pyodbc.connect(cnxn_str)

        cursor = my_cnxn.cursor()
        book = openpyxl.load_workbook(Ellicium.DBConfigurations.Config_file.Excel_Report_File)
        sheet = book.active
        cell = sheet.cell(row=5, column=3)

        cursor.execute(cell.value)
        db_list4 = []

        for i in cursor:
            db_list4.append(i)


        #print("list1",db_list1)

        return db_list4

    """def Qurry4(self):

        cnxn_str = (connection_string)

        my_cnxn = pyodbc.connect(cnxn_str)
        row = [1, 2, 3]
        cursor = my_cnxn.cursor()
        book = openpyxl.load_workbook(excelInputQuerry)
        sheet = book.active

        db_list1 = []
        db_list2 = []
        db_list3 = []

        for i in row:
            cell = sheet.cell(row=i, column=1)
            cursor.execute(cell.value)
            for ic in cursor:
                if i == 1:
                    db_list1.append(ic)
                if i == 2:
                    db_list2.append(ic)
                if i == 3:
                    db_list3.append(ic)
        #list1=db_list1
        #list2=db_list2
        #list3=db_list3



        #print("db_list1:", list1)
        #print("db_list2:", list2)
        #print("db_list3:", list3)"""